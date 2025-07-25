import streamlit as st
import pandas as pd
import pdfplumber
import io
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import coordinate_to_tuple
import zipfile

# =============================
# PDF Extraction — Clean Fund Name + Status
# =============================
def extract_funds_from_pdf(pdf_file):
    fund_data = []
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if "Fund Scorecard" not in text or "Criteria Threshold" in text:
                continue

            lines = text.split("\n")
            for i, line in enumerate(lines):
                if "Manager Tenure" in line and i > 0:
                    fund_name_candidate = lines[i - 1].strip()

                    status = None
                    if "Meets Watchlist Criteria" in fund_name_candidate:
                        fund_name = fund_name_candidate.replace("Fund Meets Watchlist Criteria.", "").strip()
                        status = "Pass"
                    elif "placed on watchlist" in fund_name_candidate:
                        fund_name = fund_name_candidate.split(" Fund has been placed")[0].strip()
                        status = "Review"
                    else:
                        fund_name = fund_name_candidate.strip()

                    if fund_name and status:
                        fund_data.append((fund_name, status))
    return fund_data

# =============================
# Excel Matching + Coloring — Fill Only, No Text
# =============================
def update_excel(excel_file, sheet_name, fund_data, investment_options, status_cell, threshold):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    try:
        start_row, col_index = coordinate_to_tuple(status_cell)
    except Exception:
        raise ValueError("Invalid cell reference for status cell.")

    fund_dict = {str(name).strip(): str(status).strip()
                 for name, status in fund_data if isinstance((name, status), tuple)}

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    results = []
    for i, fund in enumerate(investment_options):
        fund = fund.strip()
        if not fund:
            continue

        match_result = process.extractOne(fund, fund_dict.keys(), scorer=fuzz.token_sort_ratio)
        best_match = match_result[0] if match_result else None
        score = match_result[1] if match_result and len(match_result) > 1 else 0

        status = fund_dict.get(best_match, "") if score >= threshold else ""

        cell = ws.cell(row=start_row + i, column=col_index)
        cell.value = None  # Clear formula or weird characters

        if score >= threshold:
            if status == "Pass":
                cell.fill = green
            elif status == "Review":
                cell.fill = red
        else:
            cell.fill = PatternFill(fill_type=None)

        results.append({
            "Your Input": fund,
            "Matched Fund": best_match or "",
            "Status": status or "",
            "Match Score": round(score, 1)
        })

    return wb, results

# =============================
# Check for External Links
# =============================
def has_external_links(xlsx_file):
    try:
        with zipfile.ZipFile(xlsx_file) as zf:
            return any(name.startswith("xl/externalLinks/") for name in zf.namelist())
    except:
        return False

# =============================
# Streamlit App
# =============================
def run():
    st.title("Fund Scorecard")

    with st.expander("How to Use This Tool"):
        st.markdown("""
        1. **Upload a MPI PDF** — This is where fund statuses are extracted from.  
        2. **Upload an Excel Template** — This is the file where statuses will be filled in.  
        3. **Paste Investment Options** — One fund name per line, in the same order as they appear in Excel.  
        4. **Enter the cell where 'Current Quarter Status' begins** — e.g., `L6`  
        5. **Adjust the Match Score Threshold** if needed — Higher means stricter matching.  
        6. **Click Run Matching** — View matches, download updated Excel, and export the match summary.
        """)

    pdf_file = st.file_uploader("Upload MPI PDF", type="pdf")
    excel_file = st.file_uploader("Upload Excel File", type="xlsx")

    if excel_file and has_external_links(excel_file):
        st.markdown("""
        <div style='background-color:#dff0d8; padding:15px; border-radius:8px;'>
            <strong> Notice About Linked Excel Files</strong><br><br>
            This file contains <strong>external references</strong> to other workbooks (e.g., formulas linked to another Excel file).<br><br>
            When you download the updated version, Excel will display warnings like:<br>
            • “We found a problem with some content...”<br>
            • “Do you want us to try to recover...”<br><br>
             This is <strong>normal</strong>. Just click <strong>Yes</strong> and then <strong>Enable Editing</strong> when prompted — your file will open correctly.
        </div>
        """, unsafe_allow_html=True)

    investment_input = st.text_area("Paste Investment Options (one per line):")
    investment_options = [line.strip() for line in investment_input.split("\n") if line.strip()]

    match_threshold = st.slider("Minimum Match Score (fuzzy logic)", 0, 100, 20, step=5)

    if excel_file:
        xls = pd.ExcelFile(excel_file)
        sheet_name = st.selectbox("Select Excel Sheet", xls.sheet_names)
    else:
        sheet_name = None

    status_cell = st.text_input("Enter starting cell for 'Current Quarter Status' column (e.g. L6)")

    if st.button("Run Matching"):
        if not pdf_file or not excel_file or not investment_options or not status_cell or not sheet_name:
            st.error("Please provide all required inputs.")
            return

        try:
            fund_data = extract_funds_from_pdf(pdf_file)
            if not fund_data:
                st.warning("No funds extracted from PDF.")
                return

            wb, match_results = update_excel(excel_file, sheet_name, fund_data, investment_options, status_cell, match_threshold)

            st.subheader("Match Preview")
            df_results = pd.DataFrame(match_results)
            st.dataframe(df_results)

            low_conf = df_results[df_results["Match Score"] < match_threshold]
            if not low_conf.empty:
                with st.expander("⚠️ Low Confidence Matches (below threshold)"):
                    st.dataframe(low_conf)

            # Download buttons
            st.download_button("Download Match Summary CSV", df_results.to_csv(index=False), file_name="Fund_Match_Results.csv")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.success("Excel updated successfully.")
            st.download_button("Download Updated Excel", data=output, file_name="Updated_Fund_Scorecard.xlsx")

        except Exception as e:
            st.error(f"❌ Failed to update Excel: {str(e)}")

    st.markdown("---")
    st.caption("This content was generated using automation and may not be perfectly accurate. Please verify against official sources.")
